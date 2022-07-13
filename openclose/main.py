
from openpyxl import Workbook, load_workbook


def open_workbook(path):
    try:
        workbook = load_workbook(path)
        return workbook
    except FileNotFoundError:
        print("No existe un Excel con ese nombre: {}".format(FileNotFoundError))


def create_workbook():
    pass 



def config_sheet(active_workbook, sheet_title, first_title, second_title, third_title, fourth_title):
    product_sheet = active_workbook.active #accedemos a la hoja actual, en este caso la primera


    #asignamos un nombre a esa hoja
    product_sheet.title = sheet_title

    #asignamos los títulos, van en la primera fila de la hoja de trabajo
    product_sheet.cell(row=1, column=1, value = first_title )
    product_sheet.cell(row=1, column=2, value = second_title )
    product_sheet.cell(row=1, column=3, value = third_title )
    product_sheet.cell(row=1, column=4, value = fourth_title )




def set_data(active_workbook, data_to_save):
    product_sheet = active_workbook.active
    product_sheet.append(data_to_save)







if __name__ == "__main__":

    workbook = open_workbook("productors.xlsx") 
    if workbook:
        print("Libro activo {}".format(workbook))
        config_sheet(workbook,"Hoja de productos","Código de país", "Código empresa", "Código producto", "Dígito de control")

        #La lógica para el ingreso de los datos
        ingresa = True
        datos_ingreso = []
        while(ingresa):
            cp = int(input("Ingresa el códigp del pais:\t"))
            datos_ingreso.append(cp)
            ce =  int(input("Ingresa el códigp de la empresa:\t"))
            datos_ingreso.append(ce)
            cpp = int(input("Ingresa el códigp del producto:\t"))
            datos_ingreso.append(cpp)
            cdc = int(input("Ingresa el códigp del control:\t")) 
            datos_ingreso.append(cdc)
            
            # esta función guarda los datos
            set_data(workbook, datos_ingreso)
            
            #current_sheet.append(datos_ingreso)
            datos_ingreso.clear()

            rt = int(input("Deseas ingresar más datos 1=sí, 2=no \t "))

            if rt == 1:
                ingresa = True
            else:
                nombre_libro = "productos"
                workbook.save(nombre_libro+".xlsx")
                ingresa = False
    else:
        print("No activo, no hay libro")
        print("¿Desea crear un libro de trabajo?")
        user_reply = int(input("Sí -->  1\n No --> 2\t"))

        
        if user_reply == 1:
            config_sheet(workbook,"Hoja de productos","Código de país", "Código empresa", "Código producto", "Dígito de control")
            ingresa = True
            datos_ingreso = []
            while(ingresa):
                cp = int(input("Ingresa el códigp del pais:\t"))
                datos_ingreso.append(cp)
                ce =  int(input("Ingresa el códigp de la empresa:\t"))
                datos_ingreso.append(ce)
                cpp = int(input("Ingresa el códigp del producto:\t"))
                datos_ingreso.append(cpp)
                cdc = int(input("Ingresa el códigp del control:\t")) 
                datos_ingreso.append(cdc)
            
                 # esta función guarda los datos
                set_data(workbook, datos_ingreso)
            
                #current_sheet.append(datos_ingreso)
                datos_ingreso.clear()

                rt = int(input("Deseas ingresar más datos 1=sí, 2=no \t "))

                if rt == 1:
                    ingresa = True
                else:
                    nombre_libro = "productos"
                    workbook.save(nombre_libro+".xlsx")
                    ingresa = False
    
    print("Libro guardado-->")












    """workbook = Workbook() #creamos el libro de trabajo
    current_sheet = workbook.active #accedemos a la hoja 

    #asignamos un nombre a esa hoja
    current_sheet.title = "Productos escaneados"



    titulo1 = "Código de país"
    titulo2 = "Código empresa" 
    titulo3 = "Código producto" 
    titulo4 = "Dígito de control"

    current_sheet.cell(row=1, column=1, value = titulo1 )
    current_sheet.cell(row=1, column=2, value = titulo2 )
    current_sheet.cell(row=1, column=3, value = titulo3 )
    current_sheet.cell(row=1, column=4, value = titulo4 )


    
    ingresa = True
    datos_ingreso = []
    while(ingresa):
        cp = int(input("Ingresa el códigp del pais:\t"))
        datos_ingreso.append(cp)
        ce =  int(input("Ingresa el códigp de la empresa:\t"))
        datos_ingreso.append(ce)
        cpp = int(input("Ingresa el códigp del producto:\t"))
        datos_ingreso.append(cpp)
        cdc = int(input("Ingresa el códigp del control:\t")) 
        datos_ingreso.append(cdc)
        current_sheet.append(datos_ingreso)
        datos_ingreso.clear()

        rt = int(input("Deseas ingresar más datos 1=sí, 2=no \t "))

        if rt == 1:
            ingresa = True
        else:
            ingresa = False



    nombre_libro = "productos"
    print("Libro guardado-->",workbook.save(nombre_libro+".xlsx"))"""